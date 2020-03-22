import os
import datetime
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook, load_workbook
import sys

def get_html(airport_code, airport_path, month, file_path):
    url = BASE_URL.format(airport_code, month)
    # Early exist - already file download
    if os.path.exists(file_path):
        return 1
    if not os.path.exists(airport_path):
        os.makedirs(airport_path)
    BROWSER.get(url)
    try:
        WebDriverWait(BROWSER, 100).until(EC.presence_of_element_located((By.XPATH, XPATH)))
    except:
        print("Not availiable for {}".format(airport_code))
        return 0
    html = BROWSER.page_source
    with open(file_path, "w") as html_output:
        html_output.write(html)
    return 1


def export_data(airport_code, row, file_path, month):
    iata_code = SHEET_DATA['A'+str(row)].value
    type = SHEET_DATA['B'+str(row)].value
    name = SHEET_DATA['C'+str(row)].value
    elevation_ft = SHEET_DATA['D'+str(row)].value
    continent = SHEET_DATA['E'+str(row)].value
    iso_country = SHEET_DATA['F'+str(row)].value
    iso_region = SHEET_DATA['G'+str(row)].value
    coordinates = SHEET_DATA['J'+str(row)].value
    soup = BeautifulSoup(open(file_path), 'lxml')
    table = soup.find_all("table", "days ng-star-inserted")[0]
    reset = globals()['OUTPUT_START_ROW']-1
    print(reset, month, globals()['OUTPUT_START_ROW'])
    inbound_row = 1
    for index, table_content in enumerate(table.select("table > tbody > tr > td"), 1):
        globals()['OUTPUT_START_ROW'] = reset
        for data_loc, table_data in enumerate(table_content.select("tr > td"), 1):
            if index == 1:
                # Time
                if data_loc > 1:
                    SHEET_OUTPUT['A' + str(OUTPUT_START_ROW)].value = iata_code
                    SHEET_OUTPUT['B' + str(OUTPUT_START_ROW)].value = airport_code
                    SHEET_OUTPUT['C' + str(OUTPUT_START_ROW)].value = type
                    SHEET_OUTPUT['D' + str(OUTPUT_START_ROW)].value = name
                    SHEET_OUTPUT['E' + str(OUTPUT_START_ROW)].value = elevation_ft
                    SHEET_OUTPUT['F' + str(OUTPUT_START_ROW)].value = continent
                    SHEET_OUTPUT['G' + str(OUTPUT_START_ROW)].value = iso_country
                    SHEET_OUTPUT['H' + str(OUTPUT_START_ROW)].value = iso_region
                    SHEET_OUTPUT['I' + str(OUTPUT_START_ROW)].value = coordinates
                    SHEET_OUTPUT['J' + str(OUTPUT_START_ROW)].value = datetime.date(2020, month, int(table_data.get_text()))
                    globals()['OUTPUT_START_ROW'] += 1
            elif index == 2:
                # Temperature (° F)
                if data_loc > 3:
                    if inbound_row == 1:
                        SHEET_OUTPUT['K' + str(OUTPUT_START_ROW)].value = float(table_data.get_text())
                        inbound_row += 1
                    elif inbound_row == 2:
                        SHEET_OUTPUT['L' + str(OUTPUT_START_ROW)].value = float(table_data.get_text())
                        inbound_row += 1
                    elif inbound_row == 3:
                        SHEET_OUTPUT['M' + str(OUTPUT_START_ROW)].value = float(table_data.get_text())
                        globals()['OUTPUT_START_ROW'] += 1
                        inbound_row = 1
            elif index == 3:
                # Dew Point (° F)
                if data_loc > 3:
                    if inbound_row == 1:
                        SHEET_OUTPUT['N' + str(OUTPUT_START_ROW)].value = float(table_data.get_text())
                        inbound_row += 1
                    elif inbound_row == 2:
                        SHEET_OUTPUT['O' + str(OUTPUT_START_ROW)].value = float(table_data.get_text())
                        inbound_row += 1
                    elif inbound_row == 3:
                        SHEET_OUTPUT['P' + str(OUTPUT_START_ROW)].value = float(table_data.get_text())
                        globals()['OUTPUT_START_ROW'] += 1
                        inbound_row = 1
            elif index == 4:
                # Humidity (%)
                if data_loc > 3:
                    if inbound_row == 1:
                        SHEET_OUTPUT['Q' + str(OUTPUT_START_ROW)].value = float(table_data.get_text())
                        inbound_row += 1
                    elif inbound_row == 2:
                        SHEET_OUTPUT['R' + str(OUTPUT_START_ROW)].value = float(table_data.get_text())
                        inbound_row += 1
                    elif inbound_row == 3:
                        SHEET_OUTPUT['S' + str(OUTPUT_START_ROW)].value = float(table_data.get_text())
                        globals()['OUTPUT_START_ROW'] += 1
                        inbound_row = 1
            elif index == 5:
                # Wind Speed (mph)
                if data_loc > 3:
                    if inbound_row == 1:
                        SHEET_OUTPUT['T' + str(OUTPUT_START_ROW)].value = float(table_data.get_text())
                        inbound_row += 1
                    elif inbound_row == 2:
                        SHEET_OUTPUT['U' + str(OUTPUT_START_ROW)].value = float(table_data.get_text())
                        inbound_row += 1
                    elif inbound_row == 3:
                        SHEET_OUTPUT['V' + str(OUTPUT_START_ROW)].value = float(table_data.get_text())
                        globals()['OUTPUT_START_ROW'] += 1
                        inbound_row = 1
            elif index == 6:
                # Pressure (Hg)
                if data_loc > 3:
                    if inbound_row == 1:
                        SHEET_OUTPUT['W' + str(OUTPUT_START_ROW)].value = float(table_data.get_text())
                        inbound_row += 1
                    elif inbound_row == 2:
                        SHEET_OUTPUT['X' + str(OUTPUT_START_ROW)].value = float(table_data.get_text())
                        inbound_row += 1
                    elif inbound_row == 3:
                        SHEET_OUTPUT['Y' + str(OUTPUT_START_ROW)].value = float(table_data.get_text())
                        globals()['OUTPUT_START_ROW'] += 1
                        inbound_row = 1
            elif index == 7:
                # Precipitation (in)
                if data_loc > 1:
                        SHEET_OUTPUT['Z' + str(OUTPUT_START_ROW)].value = float(table_data.get_text())
                        globals()['OUTPUT_START_ROW'] += 1


if __name__ == "__main__":
    global BROWSER
    global BASE_HTML_PATH
    global BASE_URL
    global XPATH
    global SHEET_DATA
    global SHEET_OUTPUT
    global OUTPUT_START_ROW
    XPATH = '//*[@id="inner-content"]/div[2]/div[1]/div[5]/div[1]/div/lib-city-history-observation/div/div[2]/table'
    BASE_HTML_PATH = os.path.join(os.getcwd(), 'HTML')
    BASE_URL = "https://www.wunderground.com/history/monthly/{}/date/2020-{}"
    chrome_options = Options()
    chrome_options.add_argument("--headless")
    chrome_options.add_argument("--window-size=1920x1080")
    driver_path = os.path.join(os.getcwd(), "Driver")
    chrome_driver = os.path.join(driver_path, "chromedriver")
    BROWSER = webdriver.Chrome(options=chrome_options, executable_path=chrome_driver)
    start_row = 2
    end_row = 2
    iata_data = load_workbook(os.path.join(os.getcwd(), "IATA.xlsx"))
    SHEET_DATA = iata_data["IATA"]
    output_path = os.path.join(os.path.join(os.getcwd(), 'Output'), '{}_{}.xlsx'.format(start_row, end_row))
    excel_output = Workbook()
    excel_output.create_sheet('data')
    SHEET_OUTPUT = excel_output['data']
    OUTPUT_START_ROW = 3
    for row in range(start_row, end_row+1):
        airport_code = SHEET_DATA['I'+str(row)].value
        airport_path = os.path.join(BASE_HTML_PATH, airport_code)
        for month in range(1, 4):
            file_path = "{}/{}.html".format(airport_path, month)
            status_code = get_html(airport_code, airport_path, month, file_path)
            if not status_code:
                break
            export_data(airport_code, row, file_path, month)
    iata_data.close()
    excel_output.save(output_path)
    BROWSER.quit()

